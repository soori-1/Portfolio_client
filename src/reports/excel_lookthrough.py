"""
Write look-through results to a formatted Excel workbook.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Brand colours
MAROON    = "8B1A1A"
GOLD      = "C9A84C"
CREAM     = "F5ECD7"
WHITE     = "FFFFFF"
LIGHT_ROW = "FAF3E8"


def write_lookthrough_excel(
    sheets: dict,
    output_path: Path,
    portfolio_name: str = "Momentum Global ETF Portfolio",
    as_of: str = "",
) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    summary_df, top10_df = sheets["Concentration"]

    _write_stock_sheet(wb, sheets["Stock Look-Through"], portfolio_name, as_of)
    _write_rollup_sheet(wb, sheets["Sector Exposure"],  "Sector Exposure",  "Sector")
    _write_rollup_sheet(wb, sheets["Country Exposure"], "Country Exposure", "Country")
    _write_concentration_sheet(wb, summary_df, top10_df)
    _write_etf_summary_sheet(wb, sheets["ETF Summary"])
    _write_quality_sheet(wb, sheets["Data Quality"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ── helpers ────────────────────────────────────────────────────────────────

def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", start_color=color, end_color=color)


def _style_header_row(ws, n_cols: int, color: str = MAROON) -> None:
    bold_white = Font(name="Arial", bold=True, color=WHITE, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _header_fill(color)
        cell.font = bold_white
        cell.alignment = center


def _style_data_rows(ws, start_row: int, n_cols: int) -> None:
    normal = Font(name="Arial", size=9)
    alt_fill = PatternFill("solid", start_color=LIGHT_ROW, end_color=LIGHT_ROW)
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_col=n_cols), start=0):
        for cell in row:
            cell.font = normal
            if row_idx % 2 == 1:
                cell.fill = alt_fill


def _auto_width(ws, min_w: int = 8, max_w: int = 50) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        width = max(
            (len(str(cell.value)) if cell.value else 0) for cell in col
        )
        ws.column_dimensions[col_letter].width = max(min_w, min(width + 2, max_w))


def _write_stock_sheet(wb: Workbook, df: pd.DataFrame, portfolio_name: str, as_of: str) -> None:
    ws = wb.create_sheet("Stock Look-Through")
    ws.freeze_panes = "A2"

    # Title row
    ws.append([f"{portfolio_name} — Stock Look-Through   |   As of: {as_of}"])
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color=MAROON)
    ws.insert_rows(2)

    # Header
    cols = ["Rank", "Ticker", "Security Name", "Sector", "Country",
            "Portfolio Weight (%)", "Contributing ETFs", "ETF Count"]
    ws.append(cols)
    _style_header_row(ws, len(cols))

    # Data
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        ws.append([
            i,
            row["Ticker"],
            row["Security Name"],
            row["Sector"],
            row["Country"],
            row["Portfolio Weight (%)"],
            row["Contributing ETFs"],
            row["ETF Count"],
        ])

    _style_data_rows(ws, start_row=4, n_cols=len(cols))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 10

    # Format weight column as percentage-style number
    for row in ws.iter_rows(min_row=4, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = "0.0000"


def _write_rollup_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str, group_col: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"

    cols = [group_col, "Portfolio Weight (%)", "% of Classified"]
    ws.append(cols)
    _style_header_row(ws, len(cols))

    for _, row in df.iterrows():
        ws.append([row[group_col], row["Portfolio Weight (%)"], row["% of Classified"]])

    _style_data_rows(ws, start_row=2, n_cols=len(cols))
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "0.00"


def _write_concentration_sheet(wb: Workbook, summary: pd.DataFrame, top10: pd.DataFrame) -> None:
    ws = wb.create_sheet("Concentration")

    # Summary block
    ws.append(["Concentration Metrics"])
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color=MAROON)
    ws.append([])

    for _, row in summary.iterrows():
        ws.append([row["Metric"], row["Value"]])
        ws.cell(row=ws.max_row, column=1).font = Font(name="Arial", bold=True, size=9)
        ws.cell(row=ws.max_row, column=2).font = Font(name="Arial", size=9)

    ws.append([])
    ws.append(["Top 10 Holdings"])
    ws.cell(row=ws.max_row, column=1).font = Font(name="Arial", bold=True, size=11, color=MAROON)
    ws.append([])

    top10_start = ws.max_row + 1
    cols = list(top10.columns)
    ws.append(cols)
    _style_header_row(ws, len(cols))
    for _, row in top10.iterrows():
        ws.append(list(row))
    _style_data_rows(ws, start_row=top10_start + 1, n_cols=len(cols))

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 20


def _write_etf_summary_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("ETF Summary")
    ws.freeze_panes = "A2"
    cols = list(df.columns)
    ws.append(cols)
    _style_header_row(ws, len(cols))
    for _, row in df.iterrows():
        ws.append(list(row))
    _style_data_rows(ws, start_row=2, n_cols=len(cols))
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 60


def _write_quality_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Data Quality")
    ws.freeze_panes = "A2"
    cols = list(df.columns)
    ws.append(cols)
    _style_header_row(ws, len(cols), color=GOLD)
    for _, row in df.iterrows():
        ws.append(list(row))
    _style_data_rows(ws, start_row=2, n_cols=len(cols))
    _auto_width(ws)
