"""
Build initial config files for the RH Portfolio Reporting Platform.

Run once during setup. Generates:
  data/config/portfolio_weights.xlsx
  data/config/etf_sources.xlsx
  data/config/etf_classification.xlsx
  data/config/sector_overrides.xlsx
  data/config/country_overrides.xlsx
  data/config/ticker_aliases.xlsx
  data/config/report_config.xlsx
  data/config/disclaimer.txt

After running, you can edit these in Excel any time.
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
CONFIG = ROOT / "data" / "config"
CONFIG.mkdir(parents=True, exist_ok=True)

# ------------------------------------------------------------------
# 1. Portfolio weights (the 13-ETF Momentum Global Fund)
# ------------------------------------------------------------------
PORTFOLIO = [
    ("EEM",  "iShares MSCI Emerging Markets ETF",                  15.0),
    ("VT",   "Vanguard Total World Stock ETF",                     15.0),
    ("ARTY", "Tuttle Capital AI ETF",                               10.0),
    ("IYZ",  "iShares U.S. Telecommunications ETF",                 7.0),
    ("EPOL", "iShares MSCI Poland ETF",                              7.0),
    ("EWT",  "iShares MSCI Taiwan ETF",                              7.0),
    ("EWZ",  "iShares MSCI Brazil ETF",                              7.0),
    ("EWA",  "iShares MSCI Australia ETF",                           7.0),
    ("MCHI", "iShares MSCI China ETF",                               5.0),
    ("DTCR", "Global X Data Center & Digital Infrastructure ETF",    5.0),
    ("LIT",  "Global X Lithium & Battery Tech ETF",                  5.0),
    ("HYDR", "Global X Hydrogen ETF",                                5.0),
    ("IAU",  "iShares Gold Trust",                                   5.0),
]

# ------------------------------------------------------------------
# 2. ETF sources (issuer + URL pattern)
# ------------------------------------------------------------------
# Note: ARTY is Tuttle Capital, not iShares — needs custom handling or manual.
SOURCES = [
    ("EEM",  "iShares",  "https://www.ishares.com/us/products/239637/ishares-msci-emerging-markets-etf"),
    ("VT",   "Vanguard", "https://investor.vanguard.com/etf/profile/portfolio/vt"),
    ("ARTY", "Manual",   ""),  # Tuttle Capital — manual upload until we have a fetcher
    ("IYZ",  "iShares",  "https://www.ishares.com/us/products/239509/ishares-us-telecommunications-etf"),
    ("EPOL", "iShares",  "https://www.ishares.com/us/products/239668/ishares-msci-poland-etf"),
    ("EWT",  "iShares",  "https://www.ishares.com/us/products/239697/ishares-msci-taiwan-etf"),
    ("EWZ",  "iShares",  "https://www.ishares.com/us/products/239612/ishares-msci-brazil-etf"),
    ("EWA",  "iShares",  "https://www.ishares.com/us/products/239607/ishares-msci-australia-etf"),
    ("MCHI", "iShares",  "https://www.ishares.com/us/products/239619/ishares-msci-china-etf"),
    ("DTCR", "GlobalX",  "https://www.globalxetfs.com/funds/dtcr/"),
    ("LIT",  "GlobalX",  "https://www.globalxetfs.com/funds/lit/"),
    ("HYDR", "GlobalX",  "https://www.globalxetfs.com/funds/hydr/"),
    ("IAU",  "iShares",  "https://www.ishares.com/us/products/239561/ishares-gold-trust-fund"),
]

# ------------------------------------------------------------------
# 3. ETF classification (for the page 2 donut charts)
# ------------------------------------------------------------------
CLASSIFICATION = [
    ("EEM",  "Emerging Markets",  "Equity"),
    ("VT",   "Developed Markets", "Equity"),
    ("ARTY", "Developed Markets", "Equity"),
    ("IYZ",  "Developed Markets", "Equity"),
    ("EPOL", "Emerging Markets",  "Equity"),
    ("EWT",  "Emerging Markets",  "Equity"),
    ("EWZ",  "Emerging Markets",  "Equity"),
    ("EWA",  "Developed Markets", "Equity"),
    ("MCHI", "Emerging Markets",  "Equity"),
    ("DTCR", "Developed Markets", "Equity"),
    ("LIT",  "Developed Markets", "Equity"),
    ("HYDR", "Developed Markets", "Equity"),
    ("IAU",  "Commodity",         "Commodity"),
]


def _style_header(ws, n_cols):
    """Apply a maroon header row."""
    maroon = PatternFill("solid", start_color="8B1A1A", end_color="8B1A1A")
    white_bold = Font(name="Arial", bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = maroon
        cell.font = white_bold
        cell.alignment = center


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def build_portfolio_weights():
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Weights"
    ws.append(["ETF Ticker", "ETF Name", "Portfolio Weight (%)"])
    for ticker, name, weight in PORTFOLIO:
        ws.append([ticker, name, weight])
    # Validation: sum to 100
    ws.append(["", "TOTAL", f"=SUM(C2:C{len(PORTFOLIO) + 1})"])
    ws.cell(row=len(PORTFOLIO) + 2, column=2).font = Font(bold=True)
    ws.cell(row=len(PORTFOLIO) + 2, column=3).font = Font(bold=True)
    _style_header(ws, 3)
    _set_widths(ws, [14, 55, 22])
    out = CONFIG / "portfolio_weights.xlsx"
    wb.save(out)
    print(f"  ✓ {out.name}")


def build_etf_sources():
    wb = Workbook()
    ws = wb.active
    ws.title = "ETF Sources"
    ws.append(["Ticker", "Issuer", "Product URL"])
    for ticker, issuer, url in SOURCES:
        ws.append([ticker, issuer, url])
    _style_header(ws, 3)
    _set_widths(ws, [12, 12, 95])
    out = CONFIG / "etf_sources.xlsx"
    wb.save(out)
    print(f"  ✓ {out.name}")


def build_etf_classification():
    wb = Workbook()
    ws = wb.active
    ws.title = "Classification"
    ws.append(["Ticker", "Market Type", "ETF Type"])
    for ticker, market, etf_type in CLASSIFICATION:
        ws.append([ticker, market, etf_type])
    _style_header(ws, 3)
    _set_widths(ws, [12, 22, 14])
    out = CONFIG / "etf_classification.xlsx"
    wb.save(out)
    print(f"  ✓ {out.name}")


def build_overrides():
    """Sector and country override files (start empty, grow over time)."""
    for fname, cols in [
        ("sector_overrides.xlsx",  ["Ticker", "Override Sector",  "Note"]),
        ("country_overrides.xlsx", ["Ticker", "Override Country", "Note"]),
        ("ticker_aliases.xlsx",    ["Source Ticker", "Canonical Ticker", "Canonical Name", "Note"]),
    ]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Overrides"
        ws.append(cols)
        _style_header(ws, len(cols))
        _set_widths(ws, [18] * len(cols))
        out = CONFIG / fname
        wb.save(out)
        print(f"  ✓ {out.name}")


def build_report_config():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Config"
    rows = [
        ("Setting", "Value"),
        ("Portfolio Name",      "Momentum Global ETF Portfolio"),
        ("Benchmark Name",      "MSCI ACWI"),
        ("Company Name",        "Right Horizons Wealth Management"),
        ("Report Tagline",      "Guiding you in achieving your goals"),
        ("Dashboard Password",  "change_me_before_deploying"),
    ]
    for r in rows:
        ws.append(r)
    _style_header(ws, 2)
    _set_widths(ws, [26, 50])
    out = CONFIG / "report_config.xlsx"
    wb.save(out)
    print(f"  ✓ {out.name}")


def build_disclaimer():
    text = (
        "This report is prepared for informational purposes only and does not "
        "constitute investment advice, a solicitation, or an offer to buy or "
        "sell any security. The performance data depicted reflects portfolio-"
        "level returns and may vary across individual accounts due to timing of "
        "investments, redemptions, and portfolio composition. Past performance "
        "is not indicative of future results, and there is no assurance or "
        "guarantee of returns. Investments in securities are subject to market "
        "risks, including the possible loss of principal. The securities and "
        "ETFs mentioned herein do not constitute a recommendation, and "
        "portfolio holdings are subject to change. All data and information "
        "used in this report is as of the date mentioned and may not be "
        "relevant later. Investors are advised to consult their financial, "
        "legal, and tax advisors before making any investment decisions. The "
        "portfolio manager takes no responsibility for updating any "
        "information contained herein."
    )
    out = CONFIG / "disclaimer.txt"
    out.write_text(text)
    print(f"  ✓ {out.name}")


if __name__ == "__main__":
    print("Building config files in:", CONFIG)
    build_portfolio_weights()
    build_etf_sources()
    build_etf_classification()
    build_overrides()
    build_report_config()
    build_disclaimer()
    print("\nDone. Edit these files in Excel whenever you need to.")
